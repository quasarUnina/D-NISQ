import logging
import math
from threading import *
from time import perf_counter
from statistics import mean
from copy import deepcopy
from typing import Optional, Dict, Union, List
from qiskit import *
from qiskit.aqua import QuantumInstance, aqua_globals
from qiskit.aqua.algorithms.amplitude_amplifiers.grover import Grover
from qiskit.providers import BaseBackend
from qiskit.providers import Backend
from qiskit.providers.aer import QasmSimulator
from qiskit.providers.aer.noise import NoiseModel
from qiskit.circuit.library import QuadraticForm
from qiskit.optimization.algorithms import GroverOptimizer
from qiskit.optimization.algorithms.optimization_algorithm import (OptimizationResultStatus, OptimizationAlgorithm,
                                                                   OptimizationResult)
from qiskit.optimization.converters.quadratic_program_to_qubo import QuadraticProgramToQubo, QuadraticProgramConverter
from qiskit.optimization.problems import Variable
from qiskit.tools.visualization import circuit_drawer, plot_histogram
from qiskit.test.mock import *
from qiskit.ignis.mitigation.measurement import complete_meas_cal, CompleteMeasFitter

from qiskit.optimization.algorithms import MinimumEigenOptimizer
from qiskit.aqua.algorithms import NumPyMinimumEigensolver
from qiskit.optimization.problems import QuadraticProgram
from docplex.mp.model import Model
from openpyxl import Workbook
import numpy as np

logger = logging.getLogger(__name__)


class DistributedGroverOptimizer(OptimizationAlgorithm):
    """Uses Grover Adaptive Search (GAS) to find the minimum of a QUBO function."""

    def __init__(self, num_value_qubits: int, num_iterations: int = 3,
                 quantum_instance: Optional[Union[BaseBackend, Backend, QuantumInstance]] = None,
                 converters: Optional[Union[QuadraticProgramConverter,
                                            List[QuadraticProgramConverter]]] = None,
                 penalty: Optional[float] = None) -> None:
        """
        Args:
            num_value_qubits: The number of value qubits.
            num_iterations: The number of iterations the algorithm will search with
                no improvement.
            quantum_instance: Instance of selected backend, defaults to Aer's statevector simulator.
            converters: The converters to use for converting a problem into a different form.
                By default, when None is specified, an internally created instance of
                :class:`~qiskit.optimization.converters.QuadraticProgramToQubo` will be used.
            penalty: The penalty factor used in the default
                :class:`~qiskit.optimization.converters.QuadraticProgramToQubo` converter

        Raises:
            TypeError: When there one of converters is an invalid type.
        """
        self._num_value_qubits = num_value_qubits
        self._num_key_qubits = None
        self._n_iterations = num_iterations
        self._quantum_instance = None

        if quantum_instance is not None:
            self.quantum_instance = quantum_instance

        self._converters = self._prepare_converters(converters, penalty)

    @property
    def quantum_instance(self) -> QuantumInstance:
        """The quantum instance to run the circuits.

        Returns:
            The quantum instance used in the algorithm.
        """
        return self._quantum_instance

    @quantum_instance.setter
    def quantum_instance(self, quantum_instance: Union[Backend,
                                                       BaseBackend, QuantumInstance]) -> None:
        """Set the quantum instance used to run the circuits.

        Args:
            quantum_instance: The quantum instance to be used in the algorithm.
        """
        if isinstance(quantum_instance, (BaseBackend, Backend)):
            self._quantum_instance = QuantumInstance(quantum_instance)
        else:
            self._quantum_instance = quantum_instance

    def get_compatibility_msg(self, problem: QuadraticProgram) -> str:
        """Checks whether a given problem can be solved with this optimizer.

       Checks whether the given problem is compatible, i.e., whether the problem can be converted
       to a QUBO, and otherwise, returns a message explaining the incompatibility.

       Args:
           problem: The optimization problem to check compatibility.

       Returns:
           A message describing the incompatibility.
       """
        return QuadraticProgramToQubo.get_compatibility_msg(problem)

    def _get_a_operator(self, qr_key_value, problem):
        quadratic = problem.objective.quadratic.to_array()
        linear = problem.objective.linear.to_array()
        offset = problem.objective.constant

        # Get circuit requirements from input.
        quadratic_form = QuadraticForm(self._num_value_qubits, quadratic, linear, offset,
                                       little_endian=False)

        a_operator = QuantumCircuit(qr_key_value)
        a_operator.h(list(range(self._num_key_qubits)))
        a_operator.compose(quadratic_form, inplace=True)
        return a_operator

    def _get_oracle(self, qr_key_value):
        # Build negative value oracle O.
        if qr_key_value is None:
            qr_key_value = QuantumRegister(self._num_key_qubits + self._num_value_qubits)

        oracle_bit = QuantumRegister(1, "oracle")
        oracle = QuantumCircuit(qr_key_value, oracle_bit)
        oracle.z(self._num_key_qubits)  # recognize negative values.

        def is_good_state(self, measurement):
            """Check whether ``measurement`` is a good state or not."""
            value = measurement[self._num_key_qubits:self._num_key_qubits + self._num_value_qubits]
            return value[0] == '1'

        return oracle, is_good_state

    def solve(self, problem: QuadraticProgram, subproblem_label, n_threads, backend, show_steps, threadlock,
              threshold_list) -> OptimizationResult:
        """Tries to solves the given problem using the grover optimizer.

        Runs the optimizer to try to solve the optimization problem. If the problem cannot be,
        converted to a QUBO, this optimizer raises an exception due to incompatibility.

        Args:
            problem: The problem to be solved.

        Returns:
            The result of the optimizer applied to the problem.

        Raises:
            AttributeError: If the quantum instance has not been set.
            QiskitOptimizationError: If the problem is incompatible with the optimizer.
        """
        if self.quantum_instance is None:
            raise AttributeError('The quantum instance or backend has not been set.')

        self._verify_compatibility(problem)

        # convert problem to QUBO
        problem_ = self._convert(problem, self._converters)
        problem_init = deepcopy(problem_)

        # convert to minimization problem
        sense = problem_.objective.sense
        if sense == problem_.objective.Sense.MAXIMIZE:
            problem_.objective.sense = problem_.objective.Sense.MINIMIZE
            problem_.objective.constant = -problem_.objective.constant
            for i, val in problem_.objective.linear.to_dict().items():
                problem_.objective.linear[i] = -val
            for (i, j), val in problem_.objective.quadratic.to_dict().items():
                problem_.objective.quadratic[i, j] = -val
        self._num_key_qubits = len(problem_.objective.linear.to_array())  # type: ignore

        # Variables for tracking the optimum.
        optimum_found = False
        optimum_key = math.inf
        optimum_value = math.inf
        threshold = 0
        n_key = len(problem_.variables)
        n_value = self._num_value_qubits

        # Variables for tracking the solutions encountered.
        num_solutions = 2 ** n_key
        keys_measured = []

        # Variables for result object.
        operation_count = {}
        iteration = 0
        grover_alg_executions = 0

        # Variables for stopping if we've hit the rotation max.
        rotations = 0
        max_rotations = int(np.ceil(100 * np.pi / 4))

        # Initialize oracle helper object.
        qr_key_value = QuantumRegister(self._num_key_qubits + self._num_value_qubits)
        orig_constant = problem_.objective.constant
        measurement = not self.quantum_instance.is_statevector
        oracle, is_good_state = self._get_oracle(qr_key_value)

        while not optimum_found:

            m = 1
            improvement_found = False

            # Get oracle O and the state preparation operator A for the current threshold.
            problem_.objective.constant = orig_constant - threshold
            a_operator = self._get_a_operator(qr_key_value, problem_)

            # Iterate until we measure a negative.
            loops_with_no_improvement = 0
            while not improvement_found:
                # Determine the number of rotations.
                loops_with_no_improvement += 1
                rotation_count = int(np.ceil(aqua_globals.random.uniform(0, m - 1)))
                rotations += rotation_count
                # Apply Grover's Algorithm to find values below the threshold.
                # TODO: Utilize Grover's incremental feature - requires changes to Grover.
                grover = Grover(oracle,
                                state_preparation=a_operator,
                                good_state=is_good_state)
                circuit = grover.construct_circuit(rotation_count, measurement=measurement)

                # Get the next outcome.
                full_qr = QuantumRegister(self._num_key_qubits + self._num_value_qubits + 1)
                outcome = self._measure(circuit, full_qr, backend)
                k = int(outcome[0:n_key], 2)
                v = outcome[n_key:n_key + n_value]
                int_v = self._bin_to_int(v, n_value) + threshold
                v = self._twos_complement(int_v, n_value)
                logger.info('Outcome: %s', outcome)
                logger.info('Value: %s = %s', v, int_v)

                # -----------------------------------------------------------------------------------------------------
                grover_alg_executions += 1
                if show_steps:
                    print(f"\n\033[1mOutcome from Thread {subproblem_label + 1}:\033[0m"
                          f"\nThis value was measured for objective sub-function: {int_v}")
                # -----------------------------------------------------------------------------------------------------

                # If the value is an improvement, we update the iteration parameters (e.g. oracle).
                if int_v < optimum_value:
                    optimum_key = k
                    optimum_value = int_v
                    logger.info('Current Optimum Key: %s', optimum_key)
                    logger.info('Current Optimum Value: %s', optimum_value)
                    if v.startswith('1'):
                        improvement_found = True

                        # ---------------------------------------------------------------------------------------------
                        if show_steps:
                            print(f"\nThread {subproblem_label + 1} found an improvement ({optimum_value}) after "
                                  f"{loops_with_no_improvement} loop(s) with no improvement")
                        # ---------------------------------------------------------------------------------------------
                        if True:
                            if n_threads == 1:
                                threshold = optimum_value
                            else:
                                threadlock.acquire()
                                threshold_list.append(optimum_value)
                                threadlock.release()
                                threadlock.acquire()
                                threshold = min(threshold_list)
                                threadlock.release()
                        # ---------------------------------------------------------------------------------------------
                else:
                    # Using Durr and Hoyer method, increase m.
                    m = int(np.ceil(min(m * 8 / 7, 2 ** (n_key / 2))))
                    logger.info('No Improvement. M: %s', m)
                    # -------------------------------------------------------------------------------------------------
                    max_m_reached = m
                    if show_steps:
                        print(f"\nThread {subproblem_label + 1} didn't find an improvement\nTrying again with m = {m}.")
                    # -------------------------------------------------------------------------------------------------

                    # Check if we've already seen this value.
                    if k not in keys_measured:
                        keys_measured.append(k)

                    # Assume the optimal if any of the stop parameters are true.
                    if loops_with_no_improvement >= self._n_iterations or \
                            len(keys_measured) == num_solutions or rotations >= max_rotations:
                        improvement_found = True
                        optimum_found = True
                        # ---------------------------------------------------------------------------------------------
                        print(f"\nThread {subproblem_label + 1} is done")
                        # ---------------------------------------------------------------------------------------------

                # Track the operation count.
                operations = circuit.count_ops()
                operation_count[iteration] = operations
                iteration += 1

                logger.info('Operation Count: %s\n', operations)
        # If the constant is 0 and we didn't find a negative, the answer is likely 0.
        if optimum_value >= 0 and orig_constant == 0:
            optimum_key = 0

        opt_x = np.array([1 if s == '1' else 0 for s in ('{0:%sb}' % n_key).format(optimum_key)])

        # Compute function value
        fval = problem_init.objective.evaluate(opt_x)
        result = OptimizationResult(x=opt_x, fval=fval, variables=problem_.variables,
                                    status=OptimizationResultStatus.SUCCESS)

        # cast binaries back to integers
        result = self._interpret(result, self._converters)

        return GroverOptimizationResult(x=result.x, fval=result.fval, variables=result.variables,
                                        operation_counts=operation_count, n_input_qubits=n_key,
                                        n_output_qubits=n_value, intermediate_fval=fval,
                                        threshold=threshold,
                                        status=self._get_feasibility_status(problem, result.x),
                                        grover_executions=grover_alg_executions, max_m=max_m_reached)

    def _measure(self, circuit: QuantumCircuit, register: QuantumRegister, backend) -> str:
        """Get probabilities from the given backend, and picks a random outcome."""
        probs = self._get_probs(circuit, register, backend)
        freq = sorted(probs.items(), key=lambda x: x[1], reverse=True)
        # Pick a random outcome.
        freq[-1] = (freq[-1][0], 1.0 - sum(x[1] for x in freq[0:len(freq) - 1]))
        idx = aqua_globals.random.choice(len(freq), 1, p=[x[1] for x in freq])[0]
        logger.info('Frequencies: %s', freq)

        return freq[idx][0]

    def _get_probs(self, qc: QuantumCircuit, qr: QuantumRegister, backend) -> Dict[str, float]:
        """Gets probabilities from a given backend."""
        # Execute job and filter results.
        result = self.quantum_instance.execute(qc)

        if self.quantum_instance.is_statevector:
            state = np.round(result.get_statevector(qc), 5)
            keys = [bin(i)[2::].rjust(int(np.log2(len(state))), '0')[::-1]
                    for i in range(0, len(state))]
            probs = [np.round(abs(a) * abs(a), 5) for a in state]
            hist = dict(zip(keys, probs))
        else:
            state = result.get_counts(qc)
            shots = self.quantum_instance.run_config.shots
            hist = {}
            for key in state:
                hist[key[::-1]] = state[key] / shots
        hist = dict(filter(lambda p: p[1] > 0, hist.items()))
        return hist

    @staticmethod
    def _twos_complement(v: int, n_bits: int) -> str:
        """Converts an integer into a binary string of n bits using two's complement."""
        # assert -2 ** n_bits <= v < 2 ** n_bits

        if v < 0:
            v += 2 ** n_bits
            bin_v = bin(v)[2:]
        else:
            format_string = '{0:0' + str(n_bits) + 'b}'
            bin_v = format_string.format(v)

        return bin_v

    @staticmethod
    def _bin_to_int(v: str, num_value_bits: int) -> int:
        """Converts a binary string of n bits using two's complement to an integer."""
        if v.startswith("1"):
            int_v = int(v, 2) - 2 ** num_value_bits
        else:
            int_v = int(v, 2)

        return int_v


class GroverOptimizationResult(OptimizationResult):
    """A result object for Grover Optimization methods."""

    def __init__(self, x: Union[List[float], np.ndarray], fval: float, variables: List[Variable],
                 operation_counts: Dict[int, Dict[str, int]], n_input_qubits: int,
                 n_output_qubits: int, intermediate_fval: float, threshold: float,
                 status: OptimizationResultStatus, grover_executions: int, max_m: int) -> None:
        """
        Constructs a result object with the specific Grover properties.

        Args:
            x: The solution of the problem
            fval: The value of the objective function of the solution
            variables: A list of variables defined in the problem
            operation_counts: The counts of each operation performed per iteration.
            n_input_qubits: The number of qubits used to represent the input.
            n_output_qubits: The number of qubits used to represent the output.
            intermediate_fval: The intermediate value of the objective function of the solution,
                that is expected to be identical with ``fval``.
            threshold: The threshold of Grover algorithm.
            status: the termination status of the optimization algorithm.
        """
        super().__init__(x, fval, variables, status, None)
        self._operation_counts = operation_counts
        self._n_input_qubits = n_input_qubits
        self._n_output_qubits = n_output_qubits
        self._intermediate_fval = intermediate_fval
        self._threshold = threshold

        self._grover_executions = grover_executions
        self._max_m = max_m

    @property
    def operation_counts(self) -> Dict[int, Dict[str, int]]:
        """Get the operation counts.

        Returns:
            The counts of each operation performed per iteration.
        """
        return self._operation_counts

    @property
    def n_input_qubits(self) -> int:
        """Getter of n_input_qubits

        Returns:
            The number of qubits used to represent the input.
        """
        return self._n_input_qubits

    @property
    def n_output_qubits(self) -> int:
        """Getter of n_output_qubits

        Returns:
            The number of qubits used to represent the output.
        """
        return self._n_output_qubits

    @property
    def intermediate_fval(self) -> float:
        """Getter of the intermediate fval

        Returns:
            The intermediate value of fval before interpret.
        """
        return self._intermediate_fval

    @property
    def threshold(self) -> float:
        """Getter of the threshold of Grover algorithm.

        Returns:
            The threshold of Grover algorithm.
        """
        return self._threshold

    @property
    def grover_executions(self) -> int:
        return self._grover_executions

    @property
    def max_m(self) -> int:
        return self._max_m


def report_to_file(filename, max_iterations_list, max_iterations, prob, av_i):
    with open(filename + ".txt", mode='a') as report_file:
        for i in range(max_iterations):
            print(f"\n\nWITH {max_iterations_list[i]} ITERATIONS:"
                  f"\nSuccess probability: {prob[i]}"
                  f"\nAverage i parameter reached: {av_i[i]}", file=report_file
                  )

        print("\n\nN_iter   S_prob  Av_i",
              file=report_file)
        for i in range(max_iterations):
            print(
                "{:2}       {:.2f}      {:.3f}     ".format(
                    max_iterations_list[i],
                    prob[i],  av_i[i]),
                file=report_file)


def report_to_excel(filename, max_iterations_list, row_number, column_titles, val_1, val_2, val_3):
    wb = Workbook()
    ws = wb.active
    for column_label in range(len(column_titles)):
        ws.cell(row=1, column=column_label + 1, value=column_titles[column_label])
    for i in range(row_number):
        ws.cell(row=i + 2, column=1, value=max_iterations_list[i])
        ws.cell(row=i + 2, column=2, value=val_1[i])
        ws.cell(row=i + 2, column=3, value=val_2[i])
        ws.cell(row=i + 2, column=4, value=val_3[i])
        

    wb.save(filename + ".xlsx")


def print_report(max_iter, max_iter_list, prob, av_i):
    for n_iter in range(max_iter):
        print(f"\n\nWITH {max_iter_list[n_iter]} ITERATIONS:"
              f"\nSuccess probability: {prob[n_iter]}"
              f"\nAverage i parameter reached: {av_i[n_iter]}")

    print("\n\nN_iter   S_prob     Av_i")
    for n_iter in range(max_iter):
        print("{:2}       {:.2f}      {:.3f}".format(
            max_iter_list[n_iter], prob[n_iter], av_i[n_iter]))


def grover_optimizer_function(n_threads, subproblem_label, subprogram, result_list, n_iterations, backend, qi,
                              show_steps, threadlock, threshold_list):
    grover_optimizer = DistributedGroverOptimizer(6, num_iterations=n_iterations, quantum_instance=qi)
    result_list[subproblem_label] = grover_optimizer.solve(subprogram, subproblem_label, n_threads, backend, show_steps,
                                                           threadlock, threshold_list)
    print(f"\n\033[1mGrover Optimizer results for subproblem {subproblem_label + 1}:\033[0m")
    print("\033[1mx = {}\033[0m".format(result_list[subproblem_label].x))
    print("\033[1mfval = {}\033[0m".format(result_list[subproblem_label].fval))


def qubo_function(obj_function, n_var, number_of_subproblems, constant_bits):
    # Main Quadratic Program
    model = Model()
    x = [model.binary_var(name=f'x{i}') for i in range(n_var)]
    model.minimize(eval(obj_function))
    main_qp = QuadraticProgram()
    main_qp.from_docplex(model)
    print("\nMAIN Q.U.B.O PROBLEM TO SOLVE:\n")
    print(main_qp.export_as_lp_string())

    # Exact result
    exact_solver = MinimumEigenOptimizer(NumPyMinimumEigensolver())
    exact_result = exact_solver.solve(main_qp)
    int_exact_result = exact_result.x.astype('int32')
    min_value = exact_result.fval
    print("\nExpected result for main problem:")
    print("x={}".format(exact_result.x))
    print("fval={}".format(exact_result.fval))

    # Qubo problem distribution
    fixed_var_list = []
    if number_of_subproblems > 1:
        for i in range(number_of_subproblems):
            string_value = np.binary_repr(i, constant_bits)
            values_list = [int(string_value[j]) for j in range(constant_bits)]
            fixed_var_list.append(values_list)

        arrays_for_fixed_values = [np.array(fixed_var_list[i]) for i in range(number_of_subproblems)]

        SUBPROGRAMS = []
        for arr in arrays_for_fixed_values:
            # Fixed variables
            for i in range(constant_bits):
                x[i] = arr[i]

            model.minimize(eval(obj_function))
            qp = QuadraticProgram()
            qp.from_docplex(model)
            SUBPROGRAMS.append(qp)
            print(f"\nSub-Q.U.B.O. problem:\n")
            print(qp.export_as_lp_string())

    if number_of_subproblems == 1:
        return main_qp, exact_result, int_exact_result, min_value
    else:
        return main_qp, exact_result, int_exact_result, min_value, arrays_for_fixed_values, SUBPROGRAMS